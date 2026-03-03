"""
Data migration wizard for importing from existing ODS/XLSX workbooks.
Handles: date serials, name mapping, category normalization.
"""
import os
import csv
import json
import uuid
import re
from datetime import datetime, timedelta
from io import StringIO

# Try to import optional libs
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    HAS_ODF = True
except ImportError:
    HAS_ODF = False

from database import get_connection

# Excel epoch: days since 1900-01-00 (with leap year bug)
EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_date(serial) -> str:
    """Convert Excel date serial number to YYYY-MM-DD string."""
    try:
        n = int(float(serial))
        if 30000 <= n <= 60000:  # reasonable Excel date range
            d = EXCEL_EPOCH + timedelta(days=n)
            return d.strftime('%Y-%m-%d')
    except Exception:
        pass
    return str(serial)


def parse_flexible_date(value) -> str:
    """Parse date from various formats into YYYY-MM-DD."""
    if value is None or value == '':
        return ''
    # Already a datetime
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    # Date object
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s:
        return ''
    # Check if it's an Excel serial
    try:
        n = float(s)
        if 30000 <= n <= 60000:
            return excel_serial_to_date(int(n))
    except ValueError:
        pass
    # Try common date formats
    formats = [
        '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y',
        '%d-%b-%Y', '%B %d, %Y', '%m-%d-%Y',
        '%Y/%m/%d', '%d/%m/%Y'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s  # Return as-is if unparseable


def parse_amount(value) -> float:
    """Parse a dollar amount from various formats."""
    if value is None or value == '':
        return 0.0
    s = str(value).strip().replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def read_xlsx_sheets(filepath: str) -> dict:
    """Read all sheets from XLSX file. Returns {sheet_name: [rows]}"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h or '').strip() for h in row]
                continue
            if headers and any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))
        result[name] = {'headers': headers or [], 'rows': rows}
    return result


def read_ods_sheets(filepath: str) -> dict:
    """Read all sheets from ODS file. Returns {sheet_name: {headers, rows}}"""
    if not HAS_ODF:
        raise ImportError("odfpy not installed")
    doc = load_ods(filepath)
    result = {}
    sheets = doc.spreadsheet.getElementsByType(Table)
    for sheet in sheets:
        name = sheet.getAttribute('name')
        all_rows = []
        trows = sheet.getElementsByType(TableRow)
        for trow in trows:
            cells = trow.getElementsByType(TableCell)
            row_data = []
            for cell in cells:
                repeat = cell.getAttribute('number-columns-repeated')
                repeat = int(repeat) if repeat else 1
                # Get text content
                paragraphs = cell.getElementsByType(P)
                val = ''
                if paragraphs:
                    for p in paragraphs:
                        for node in p.childNodes:
                            if hasattr(node, 'data'):
                                val += node.data
                for _ in range(repeat):
                    row_data.append(val)
            all_rows.append(row_data)
        if not all_rows:
            continue
        headers = [str(h).strip() for h in all_rows[0]]
        rows = []
        for raw in all_rows[1:]:
            if any(v for v in raw):
                rows.append(dict(zip(headers, raw + [''] * (len(headers) - len(raw)))))
        result[name] = {'headers': headers, 'rows': rows}
    return result


# ============================================================
# SHEET → TABLE MAPPERS
# ============================================================

SHEET_MAPPINGS = {
    'Ledger': 'ledger',
    'ledger': 'ledger',
    'Time': 'timesheet',
    'Timesheet': 'timesheet',
    'Clients': 'clients',
    'Employees': 'employees',
    'Rate': 'employee_rates',
    'Contractors': 'contractors',
    'Jobs': 'jobs',
    'InvoicesSummary': 'invoices',
    'Invoices': 'invoices',
}

LEDGER_COLUMN_MAP = {
    'Date': 'entry_date',
    'date': 'entry_date',
    'Job': 'job_code',
    'Job Code': 'job_code',
    'Job #': 'job_number',
    'Invoice': 'invoice_number',
    'Invoice #': 'invoice_number',
    'Status': 'status',
    'Category': 'category',
    'Description': 'description',
    'Vendor': 'vendor',
    'COGS': 'is_cogs',
    'Amount': 'amount',
    'Receipt': 'receipt_filename',
    'Notes': 'notes',
}

TIMESHEET_COLUMN_MAP = {
    'Date': 'entry_date',
    'Job': 'job_code',
    'Job Code': 'job_code',
    'Invoice': 'invoice_number',
    'Employee': 'emp_id',
    'Emp ID': 'emp_id',
    'Hours': 'hours',
    'Bill Rate': 'bill_rate',
    'Cost Rate': 'cost_rate',
    'Bill Amount': 'bill_amount',
    'Cost Amount': 'cost_amount',
    'Expenses': 'expenses',
    'Description': 'description',
    'Notes': 'notes',
}

CLIENTS_COLUMN_MAP = {
    'Year': 'year_acquired',
    'Year Acquired': 'year_acquired',
    'Customer ID': 'customer_id',
    'Last Name': 'last_name',
    'Full Name': 'full_name',
    'Name': 'full_name',
    'Address': 'address',
    'City/State/Zip': 'city_state_zip',
    'Phone': 'phone1',
    'Phone 2': 'phone2',
    'Email': 'email1',
    'Status': 'status',
    'Notes': 'notes',
}

EMPLOYEE_COLUMN_MAP = {
    'ID': 'emp_id',
    'Emp ID': 'emp_id',
    'First Name': 'first_name',
    'Last Name': 'last_name',
    'Phone': 'phone',
    'Email': 'email',
    'Address': 'address',
    'Occupation': 'occupation',
    'Hire Date': 'hire_date',
    'Status': 'status',
    'Notes': 'notes',
}


def detect_sheet_mapping(sheet_name: str, headers: list) -> str:
    """Guess which database table a sheet maps to."""
    name_lower = sheet_name.lower()
    if 'ledger' in name_lower:
        return 'ledger'
    if 'time' in name_lower and 'summary' not in name_lower:
        return 'timesheet'
    if 'client' in name_lower or 'customer' in name_lower:
        return 'clients'
    if 'employee' in name_lower or 'emp' in name_lower:
        return 'employees'
    if 'rate' in name_lower:
        return 'employee_rates'
    if 'contractor' in name_lower or 'vendor' in name_lower or 'sub' in name_lower:
        return 'contractors'
    if 'invoice' in name_lower:
        return 'invoices'
    if 'job' in name_lower:
        return 'jobs'
    return None


def map_row_to_ledger(raw: dict) -> dict:
    """Map a raw row dict to ledger table fields."""
    mapped = {}
    for src, dst in LEDGER_COLUMN_MAP.items():
        if src in raw:
            mapped[dst] = raw[src]
    # Normalize fields
    if 'entry_date' in mapped:
        mapped['entry_date'] = parse_flexible_date(mapped['entry_date'])
    if 'amount' in mapped:
        mapped['amount'] = parse_amount(mapped['amount'])
    if 'is_cogs' in mapped:
        val = str(mapped['is_cogs']).lower()
        mapped['is_cogs'] = 1 if val in ('yes', 'true', '1', 'x') else 0
    return mapped


def map_row_to_timesheet(raw: dict) -> dict:
    """Map a raw row dict to timesheet table fields."""
    mapped = {}
    for src, dst in TIMESHEET_COLUMN_MAP.items():
        if src in raw:
            mapped[dst] = raw[src]
    if 'entry_date' in mapped:
        mapped['entry_date'] = parse_flexible_date(mapped['entry_date'])
    for field in ('hours', 'bill_rate', 'cost_rate', 'bill_amount', 'cost_amount', 'expenses'):
        if field in mapped:
            mapped[field] = parse_amount(mapped[field])
    return mapped


def map_row_to_clients(raw: dict) -> dict:
    mapped = {}
    for src, dst in CLIENTS_COLUMN_MAP.items():
        if src in raw:
            mapped[dst] = raw[src]
    if 'year_acquired' in mapped:
        try:
            mapped['year_acquired'] = int(mapped['year_acquired'])
        except Exception:
            mapped['year_acquired'] = None
    return mapped


def map_row_to_employees(raw: dict) -> dict:
    mapped = {}
    for src, dst in EMPLOYEE_COLUMN_MAP.items():
        if src in raw:
            mapped[dst] = raw[src]
    if 'emp_id' in mapped:
        try:
            mapped['emp_id'] = int(mapped['emp_id'])
        except Exception:
            pass
    if 'hire_date' in mapped:
        mapped['hire_date'] = parse_flexible_date(mapped['hire_date'])
    return mapped


# ============================================================
# VALIDATION
# ============================================================

def validate_ledger_rows(rows: list) -> tuple:
    """Return (valid_rows, errors) for ledger import."""
    valid, errors = [], []
    for i, row in enumerate(rows, 2):
        errs = []
        if not row.get('entry_date'):
            errs.append('Missing date')
        elif row['entry_date'] and len(row['entry_date']) != 10:
            errs.append(f"Unparseable date: {row['entry_date']}")
        if row.get('amount') is None:
            errs.append('Missing amount')
        if errs:
            errors.append({'row': i, 'data': row, 'errors': errs})
        else:
            valid.append(row)
    return valid, errors


def validate_timesheet_rows(rows: list) -> tuple:
    valid, errors = [], []
    for i, row in enumerate(rows, 2):
        errs = []
        if not row.get('entry_date'):
            errs.append('Missing date')
        if row.get('hours') is None:
            errs.append('Missing hours')
        if errs:
            errors.append({'row': i, 'data': row, 'errors': errs})
        else:
            valid.append(row)
    return valid, errors


# ============================================================
# IMPORT EXECUTION
# ============================================================

def import_clients(rows: list, conn) -> dict:
    from automations import generate_customer_id, extract_last_name
    imported, skipped = 0, 0
    for row in rows:
        if not row.get('full_name'):
            skipped += 1
            continue
        year = row.get('year_acquired') or datetime.now().year
        cid = row.get('customer_id') or generate_customer_id(row['full_name'], year, conn)
        last = row.get('last_name') or extract_last_name(row['full_name'])
        try:
            conn.execute("""
                INSERT OR IGNORE INTO clients
                (year_acquired, customer_id, last_name, full_name, address,
                 city_state_zip, phone1, phone2, email1, status, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, [
                year, cid, last, row['full_name'],
                row.get('address', ''), row.get('city_state_zip', ''),
                row.get('phone1', ''), row.get('phone2', ''),
                row.get('email1', ''), row.get('status', 'Active'),
                row.get('notes', '')
            ])
            imported += 1
        except Exception:
            skipped += 1
    return {'imported': imported, 'skipped': skipped}


def import_ledger(rows: list, conn) -> dict:
    imported, skipped = 0, 0
    for row in rows:
        if not row.get('entry_date') or not row.get('amount'):
            skipped += 1
            continue
        try:
            conn.execute("""
                INSERT INTO ledger
                (entry_date, job_code, job_number, invoice_number, status,
                 category, description, vendor, is_cogs, amount,
                 receipt_filename, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, [
                row['entry_date'], row.get('job_code', ''),
                row.get('job_number', ''), row.get('invoice_number', ''),
                row.get('status', 'Pending'), row.get('category', ''),
                row.get('description', ''), row.get('vendor', ''),
                row.get('is_cogs', 0), row['amount'],
                row.get('receipt_filename', ''), row.get('notes', '')
            ])
            imported += 1
        except Exception:
            skipped += 1
    return {'imported': imported, 'skipped': skipped}


def import_timesheet(rows: list, conn) -> dict:
    imported, skipped = 0, 0
    for row in rows:
        if not row.get('entry_date'):
            skipped += 1
            continue
        try:
            conn.execute("""
                INSERT INTO timesheet
                (entry_date, job_code, invoice_number, emp_id,
                 hours, bill_rate, cost_rate, bill_amount, cost_amount,
                 expenses, description, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, [
                row['entry_date'], row.get('job_code', ''),
                row.get('invoice_number', ''), row.get('emp_id'),
                row.get('hours', 0), row.get('bill_rate', 0),
                row.get('cost_rate', 0), row.get('bill_amount', 0),
                row.get('cost_amount', 0), row.get('expenses', 0),
                row.get('description', ''), row.get('notes', '')
            ])
            imported += 1
        except Exception:
            skipped += 1
    return {'imported': imported, 'skipped': skipped}


def import_employees(rows: list, conn) -> dict:
    imported, skipped = 0, 0
    max_id = conn.execute("SELECT MAX(emp_id) FROM employees").fetchone()[0] or 0
    for row in rows:
        if not row.get('first_name') and not row.get('last_name'):
            skipped += 1
            continue
        emp_id = row.get('emp_id')
        if not emp_id:
            max_id += 1
            emp_id = max_id
        try:
            conn.execute("""
                INSERT OR IGNORE INTO employees
                (emp_id, first_name, last_name, phone, email,
                 address, occupation, hire_date, status, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, [
                emp_id, row.get('first_name', ''), row.get('last_name', ''),
                row.get('phone', ''), row.get('email', ''),
                row.get('address', ''), row.get('occupation', ''),
                row.get('hire_date', ''), row.get('status', 'Active'),
                row.get('notes', '')
            ])
            imported += 1
        except Exception:
            skipped += 1
    return {'imported': imported, 'skipped': skipped}


def run_full_import(batch_id: str, sheets: dict, conn) -> dict:
    """Run full import from parsed sheets data."""
    summary = {}
    for sheet_name, sheet_data in sheets.items():
        target = detect_sheet_mapping(sheet_name, sheet_data.get('headers', []))
        rows = sheet_data.get('rows', [])
        if not target or not rows:
            continue
        if target == 'ledger':
            mapped = [map_row_to_ledger(r) for r in rows]
            valid, _ = validate_ledger_rows(mapped)
            result = import_ledger(valid, conn)
        elif target == 'timesheet':
            mapped = [map_row_to_timesheet(r) for r in rows]
            valid, _ = validate_timesheet_rows(mapped)
            result = import_timesheet(valid, conn)
        elif target == 'clients':
            mapped = [map_row_to_clients(r) for r in rows]
            result = import_clients(mapped, conn)
        elif target == 'employees':
            mapped = [map_row_to_employees(r) for r in rows]
            result = import_employees(mapped, conn)
        else:
            continue
        summary[sheet_name] = result
    conn.execute("""
        UPDATE import_batches SET status='Imported', summary=?, updated_at=datetime('now')
        WHERE batch_id=?
    """, [json.dumps(summary), batch_id])
    conn.commit()
    return summary


# ============================================================
# CSV BANK IMPORT (Phase 7 prep)
# ============================================================

def detect_bank_format(headers: list) -> str:
    """Detect which bank CSV format we have."""
    h = [x.lower().strip() for x in headers]
    if 'transaction date' in h and 'debit' in h and 'credit' in h:
        return 'capital_one'
    if 'transaction date' in h and 'amount' in h:
        return 'chase'
    if 'date' in h and 'amount' in h:
        return 'generic'
    return 'generic'


def normalize_bank_csv(content: str) -> list:
    """Parse a bank CSV and normalize to [{date, description, amount, type}]."""
    reader = csv.DictReader(StringIO(content))
    headers = reader.fieldnames or []
    fmt = detect_bank_format(headers)
    rows = []
    for row in reader:
        if fmt == 'capital_one':
            date = parse_flexible_date(row.get('Transaction Date', ''))
            desc = row.get('Description', '')
            debit = parse_amount(row.get('Debit', ''))
            credit = parse_amount(row.get('Credit', ''))
            if debit:
                amount, txtype = -abs(debit), 'Debit'
            else:
                amount, txtype = abs(credit), 'Credit'
        elif fmt == 'chase':
            date = parse_flexible_date(row.get('Transaction Date', row.get('Posting Date', '')))
            desc = row.get('Description', '')
            raw = parse_amount(row.get('Amount', ''))
            amount = raw
            txtype = 'Credit' if raw >= 0 else 'Debit'
        else:
            # Generic — try to find date/description/amount
            date_keys = ['Date', 'date', 'Transaction Date', 'Posted Date']
            desc_keys = ['Description', 'description', 'Merchant', 'Name']
            amt_keys = ['Amount', 'amount', 'Transaction Amount']
            date = ''
            for k in date_keys:
                if k in row:
                    date = parse_flexible_date(row[k])
                    break
            desc = ''
            for k in desc_keys:
                if k in row:
                    desc = row[k]
                    break
            raw = 0.0
            for k in amt_keys:
                if k in row:
                    raw = parse_amount(row[k])
                    break
            amount = raw
            txtype = 'Credit' if raw >= 0 else 'Debit'
        if date:
            rows.append({
                'transaction_date': date,
                'description': desc,
                'amount': amount,
                'transaction_type': txtype
            })
    return rows

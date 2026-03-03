"""
import_kbweb.py — Custom import script for KB Construction data

Reads: import_data.xlsx (the user's actual spreadsheet)
Writes: construction.db (or any path via --db)

Usage:
    python import_kbweb.py                          # auto-find DB
    python import_kbweb.py --db path/to/custom.db
    python import_kbweb.py --dry-run
    python import_kbweb.py --sheet Ledger           # import one sheet only

What gets imported:
    Clients     — 129 rows
    Contractors — 396 rows
    Ledger      — 3,006 rows  (expenses + income + credit-card payments)
    Timesheet   — 4,611 rows
"""

import sys, os, argparse, sqlite3
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  pip install openpyxl"); sys.exit(1)

XLSX = Path(__file__).parent / "import_data.xlsx"
SKIP = 3   # rows 1=col headers, 2=notes, 3=user's own sub-header — data starts row 4

# ─── tiny helpers ─────────────────────────────────────────────────────────────

def s(v, default=""):
    if v is None: return default
    if isinstance(v, float) and v == int(v): return str(int(v))
    return str(v).strip()

def f(v, default=0.0):
    try:
        if v is None or str(v).strip() == "": return default
        return float(str(v).replace(",","").replace("$","").strip())
    except: return default

def i(v, default=0):
    try: return int(float(str(v).strip()))
    except: return default

def d(v):
    """Return YYYY-MM-DD string or ''."""
    if v is None or str(v).strip() == "": return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    raw = str(v).strip()
    for fmt in ("%Y-%m-%d","%m/%d/%Y","%m/%d/%y","%d/%m/%Y"):
        try: return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except: pass
    return raw

def rows(ws):
    for row in ws.iter_rows(min_row=SKIP+1, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            yield row

def sep(title):
    print(f"\n{'─'*62}\n  {title}\n{'─'*62}")

# ─── CLIENTS ──────────────────────────────────────────────────────────────────
# Cols: full_name, last_name, customer_id, year_acquired, address,
#       city_state_zip, phone1, phone2, email1, email2, status, notes

def import_clients(conn, ws, dry, clear):
    sep("CLIENTS")
    if clear:
        conn.execute("UPDATE clients SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠  Cleared existing clients")

    ok = skip = err = 0
    for r in rows(ws):
        full_name = s(r[0])
        if not full_name:
            skip += 1; continue

        if conn.execute("SELECT id FROM clients WHERE full_name=? AND is_deleted=0",[full_name]).fetchone():
            print(f"  ↩  {full_name} — already exists, skipping")
            skip += 1; continue

        status = s(r[10]) or "Active"
        if status not in ("Active","Archived","Prospect"): status = "Active"

        # Clean phone — sometimes stored as integer
        ph1 = s(r[6])
        if ph1 and ph1.lstrip("-").isdigit() and len(ph1) == 10:
            ph1 = f"{ph1[:3]}-{ph1[3:6]}-{ph1[6:]}"

        # Strip non-breaking spaces from email
        email1 = s(r[8]).replace("\xa0","").strip()

        try:
            if not dry:
                conn.execute("""
                    INSERT INTO clients
                        (full_name,last_name,customer_id,year_acquired,
                         address,city_state_zip,phone1,phone2,
                         email1,email2,status,notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, [full_name, s(r[1]), s(r[2]),
                      i(r[3]) if r[3] else None,
                      s(r[4]), s(r[5]), ph1, s(r[7]),
                      email1, s(r[9]).replace("\xa0",""),
                      status, s(r[11])])
            print(f"  ✅  {full_name}")
            ok += 1
        except Exception as e:
            print(f"  ❌  {full_name}: {e}"); err += 1

    print(f"\n  Clients: {ok} imported | {skip} skipped | {err} errors")
    return ok, skip, err

# ─── CONTRACTORS ──────────────────────────────────────────────────────────────
# Cols: company_name, trade_type, contact_person, phone, cell,
#       email, address, Website, license_number, requires_1099,
#       rank_preference, notes

def import_contractors(conn, ws, dry, clear):
    sep("CONTRACTORS")
    if clear:
        conn.execute("UPDATE contractors SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠  Cleared existing contractors")

    ok = skip = err = 0
    for r in rows(ws):
        company = s(r[0])
        if not company:
            skip += 1; continue

        if conn.execute("SELECT id FROM contractors WHERE company_name=? AND is_deleted=0",[company]).fetchone():
            print(f"  ↩  {company} — already exists, skipping")
            skip += 1; continue

        # phone sometimes stored as raw integer
        ph  = s(r[3])
        cell = s(r[4])
        for p in [ph, cell]:
            digits = "".join(c for c in p if c.isdigit())
            if len(digits) == 10 and p.lstrip("-").isdigit():
                p = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"

        try:
            if not dry:
                conn.execute("""
                    INSERT INTO contractors
                        (company_name,trade_type,contact_person,phone,cell,
                         email,address,website,license_number,
                         requires_1099,rank_preference,notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, [company, s(r[1]), s(r[2]).strip(),
                      ph, cell, s(r[5]), s(r[6]),
                      s(r[7]),   # Website column
                      s(r[8]),
                      i(r[9]),   # requires_1099
                      i(r[10]),  # rank_preference
                      s(r[11])])
            print(f"  ✅  {company}")
            ok += 1
        except Exception as e:
            print(f"  ❌  {company}: {e}"); err += 1

    print(f"\n  Contractors: {ok} imported | {skip} skipped | {err} errors")
    return ok, skip, err

# ─── LEDGER ───────────────────────────────────────────────────────────────────
# Cols (0-indexed):
#   0  entry_date        datetime
#   1  Job-name          str   (we use as 'job_name' note)
#   2  job_code          str
#   3  invoice_number    str
#   4  status            str   (Work/Sent/Paid/Closed)
#   5  category          str
#   6  description       str
#   7  vendor            str
#   8  is_cogs           int
#   9  amount            float (expense amount)
#  10  receipt_filename  str
#  11  receipt_check     str   (e.g. "Receipt Missing!")
#  12  coi_check         str
#  13  payment_type      str   (check #, cash, credit, etc.)
#  14  notes             str
#  15  Checking_Income   float (money received from clients)
#  16  Checking_Outgo    float (credit card payments — balance transfers)
#
# Strategy:
#   • col[9] > 0          → normal expense row, amount = col[9]
#   • col[15] > 0         → income received, amount = col[15], category overridden to 'Income Received'
#   • col[16] > 0         → credit card payment/transfer, category overridden to 'Credit Card Payment'
#   • payment_type stored in notes as "check:#XXXX" when it looks like a check number
#   • receipt_check / coi_check stored as note flags if non-empty

def import_ledger(conn, ws, dry, clear):
    sep("LEDGER")
    if clear:
        conn.execute("UPDATE ledger SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠  Cleared existing ledger entries")

    ok = skip = err = 0
    income_count = outgo_count = 0

    for idx, r in enumerate(rows(ws), 4):
        entry_date = d(r[0])
        if not entry_date:
            skip += 1; continue

        exp_amt  = f(r[9])
        inc_amt  = f(r[15])
        outgo_amt= f(r[16])

        category     = s(r[5])
        description  = s(r[6])
        vendor       = s(r[7])
        job_code     = s(r[2])
        job_name     = s(r[1])
        invoice_num  = s(r[3])
        status_raw   = s(r[4])
        is_cogs      = i(r[8])
        receipt_file = s(r[10])
        rcpt_check   = s(r[11])
        coi_check    = s(r[12])
        pay_type     = s(r[13])
        notes_raw    = s(r[14])

        # Build enriched notes
        note_parts = []
        if notes_raw:   note_parts.append(notes_raw)
        if pay_type:    note_parts.append(f"payment:{pay_type}")
        if rcpt_check:  note_parts.append(f"receipt_status:{rcpt_check}")
        if coi_check:   note_parts.append(f"coi:{coi_check}")
        if job_name and job_name != job_code:
            note_parts.append(f"job_name:{job_name}")
        notes_full = " | ".join(note_parts)

        # Map status
        status_map = {"work":"Pending","sent":"Sent","paid":"Cleared","closed":"Cleared"}
        db_status = status_map.get(status_raw.lower(), "Cleared")

        def insert(amount, cat, cogs):
            if not dry:
                conn.execute("""
                    INSERT INTO ledger
                        (entry_date, job_code, invoice_number, status,
                         category, description, vendor, is_cogs,
                         amount, receipt_filename, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, [entry_date, job_code, invoice_num, db_status,
                      cat, description, vendor, cogs,
                      amount, receipt_file, notes_full])

        try:
            if exp_amt > 0:
                insert(exp_amt, category or "Miscellaneous", is_cogs)
                ok += 1
                print(f"  ✅  {entry_date}  ${exp_amt:>10,.2f}  {vendor or '—':<22}  {category}")

            if inc_amt > 0:
                # Income received from clients — store as positive, special category
                insert(inc_amt, "Income Received", 0)
                income_count += 1
                ok += 1
                print(f"  💰  {entry_date}  ${inc_amt:>10,.2f}  INCOME  {description[:40]}")

            if outgo_amt > 0:
                insert(outgo_amt, "Credit Card Payment", 0)
                outgo_count += 1
                ok += 1
                print(f"  💳  {entry_date}  ${outgo_amt:>10,.2f}  CC PAYMENT  {description[:30]}")

            if exp_amt <= 0 and inc_amt <= 0 and outgo_amt <= 0:
                # Row has date but no dollar amounts — still preserve as $0 memo if has description
                if description or category:
                    insert(0.0, category or "Memo", is_cogs)
                    ok += 1
                    print(f"  📝  {entry_date}  $0  memo  {description[:40]}")
                else:
                    skip += 1

        except Exception as e:
            print(f"  ❌  row {idx} ({entry_date}): {e}")
            err += 1

    print(f"\n  Ledger: {ok} rows imported ({income_count} income | {outgo_count} CC payments)")
    print(f"  Skipped: {skip} | Errors: {err}")
    return ok, skip, err

# ─── TIMESHEET ────────────────────────────────────────────────────────────────
# Cols (0-indexed):
#   0  entry_date
#   1  Day of the week   (ignored)
#   2  job_code
#   3  invoice_number
#   4  status
#   5  category          (work type — Electrical, GC, Framing-Labor, etc.)
#   6  description
#   7  bill_rate         (rate charged to client per hour)
#   8  cost_rate         (rate paid to worker per hour)
#   9  cost_total        (= hours × rate, pre-computed)
#  10  notes
#  11  person_label      (worker name)
#  12  hours
#  13  billable          (GC / Cash / etc.)
#  14  billed_status     (X = billed)
#  15  withholdings      (dollar amount withheld — ~15.6% of cost_total)

def import_timesheet(conn, ws, dry, clear):
    sep("TIMESHEET")
    if clear:
        conn.execute("UPDATE timesheet SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠  Cleared existing timesheet entries")

    ok = skip = err = 0
    for idx, r in enumerate(rows(ws), 4):
        entry_date   = d(r[0])
        person_label = s(r[11])
        job_code     = s(r[2])
        hours        = f(r[12])

        if not entry_date:
            skip += 1; continue
        if not person_label:
            skip += 1; continue
        if hours <= 0:
            skip += 1; continue

        bill_rate    = f(r[7])
        cost_rate    = f(r[8])
        cost_total   = f(r[9])    # pre-computed in spreadsheet
        withholding  = f(r[15])   # dollar amount withheld
        invoice_num  = s(r[3])
        status_raw   = s(r[4])
        category     = s(r[5])
        description  = s(r[6])
        billable     = s(r[13])   # GC / Cash / etc.
        billed_stat  = s(r[14])   # X = already billed
        notes_raw    = s(r[10])

        # bill_amount = hours × bill_rate
        bill_amount = round(hours * bill_rate, 2)
        # Use spreadsheet's pre-computed cost_total if available, else recalculate
        if cost_total > 0:
            cost_amount = cost_total
        else:
            cost_amount = round(hours * cost_rate, 2)

        # Build notes
        note_parts = []
        if notes_raw and str(notes_raw) != "0": note_parts.append(notes_raw)
        if billable:    note_parts.append(f"type:{billable}")
        if billed_stat: note_parts.append(f"billed:{billed_stat}")
        if withholding: note_parts.append(f"withholding:${withholding:.2f}")
        if category:    note_parts.append(f"work_type:{category}")
        notes_full = " | ".join(note_parts)

        try:
            if not dry:
                conn.execute("""
                    INSERT INTO timesheet
                        (entry_date, person_label, job_code,
                         invoice_number, hours,
                         bill_rate, cost_rate, bill_amount, cost_amount,
                         description, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, [entry_date, person_label, job_code,
                      invoice_num, hours,
                      bill_rate, cost_rate, bill_amount, cost_amount,
                      description, notes_full])
            print(f"  ✅  {entry_date}  {person_label:<18}  {job_code:<16}  {hours}h  ${cost_amount:,.2f}")
            ok += 1
        except Exception as e:
            print(f"  ❌  row {idx} ({entry_date} {person_label}): {e}")
            err += 1

    print(f"\n  Timesheet: {ok} imported | {skip} skipped | {err} errors")
    return ok, skip, err

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def find_db():
    for p in [Path("construction.db"), Path("construction_manager.db"),
              Path(os.environ.get("CONSTRUCTION_DB",""))]:
        if p and p.exists(): return p
    for p in Path(".").glob("*.db"):
        return p
    return None

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db",             default=None)
    parser.add_argument("--dry-run",        action="store_true")
    parser.add_argument("--clear-existing", action="store_true")
    parser.add_argument("--sheet",          default=None,
        help="Import only one sheet: Clients | Contractors | Ledger | Timesheet")
    args = parser.parse_args()

    xlsx = XLSX
    if not xlsx.exists():
        # Also look in same folder as this script
        xlsx = Path(__file__).parent / "import_data.xlsx"
    if not xlsx.exists():
        print(f"❌  Cannot find import_data.xlsx. Put it in the same folder as this script.")
        sys.exit(1)

    db_path = Path(args.db) if args.db else find_db()
    if not db_path or not db_path.exists():
        print("❌  Cannot find the database. Run the app once first to create it.")
        print("    Or pass: --db path/to/construction.db")
        sys.exit(1)

    print(f"\n{'='*62}")
    print(f"  KB Construction — Data Import")
    print(f"{'='*62}")
    print(f"  File  : {xlsx}")
    print(f"  DB    : {db_path}")
    print(f"  Mode  : {'DRY RUN — no changes' if args.dry_run else '⚡ LIVE IMPORT'}")
    if args.sheet: print(f"  Sheet : {args.sheet} only")
    if args.clear_existing:
        print(f"  ⚠️   --clear-existing ON — existing rows will be soft-deleted!")

    if not args.dry_run and args.clear_existing:
        if input("\nType YES to confirm: ").strip() != "YES":
            print("Aborted."); sys.exit(0)

    wb   = openpyxl.load_workbook(xlsx, data_only=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = OFF")

    totals = dict(ok=0, skip=0, err=0)

    def run(fn, sheet_name):
        if args.sheet and args.sheet.lower() != sheet_name.lower():
            return
        if sheet_name not in wb.sheetnames:
            print(f"\n⚠️  Sheet '{sheet_name}' not found — skipping"); return
        o, sk, e = fn(conn, wb[sheet_name], args.dry_run, args.clear_existing)
        totals["ok"]   += o
        totals["skip"] += sk
        totals["err"]  += e

    run(import_clients,    "Clients")
    run(import_contractors,"Contractors")
    run(import_ledger,     "Ledger")
    run(import_timesheet,  "Timesheet")

    if not args.dry_run:
        conn.commit()
        print("\n  ✅  Changes committed to database.")
    conn.close()

    print(f"\n{'='*62}")
    print(f"  DONE")
    print(f"{'='*62}")
    print(f"  ✅  Imported : {totals['ok']}")
    print(f"  ↩  Skipped  : {totals['skip']}")
    print(f"  ❌  Errors   : {totals['err']}")
    if args.dry_run:
        print(f"\n  ℹ️  Dry run complete — rerun without --dry-run to commit.")
    print()

if __name__ == "__main__":
    main()

"""
import_data.py — Import tool for Construction Business Manager

Usage:
    python import_data.py import_data.xlsx
    python import_data.py import_data.xlsx --db path/to/custom.db
    python import_data.py import_data.xlsx --dry-run        (preview only, no DB changes)
    python import_data.py import_data.xlsx --clear-existing  (wipe existing records first)

What it imports:
    ✅ Clients
    ✅ Contractors
    ✅ Ledger entries
    ✅ Timesheet entries

What it does NOT touch:
    — Invoices, payroll, estimates, employees, settings
    — Any rows already in the database (imports are additive by default)
"""

import sys
import os
import argparse
import sqlite3
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  openpyxl is not installed. Run:  pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_DB_PATHS = [
    Path("construction.db"),
    Path("construction_manager.db"),
    Path(os.environ.get("CONSTRUCTION_DB", "")),
]

SKIP_ROWS = 2   # row 1 = header, row 2 = notes


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def find_db():
    for p in DEFAULT_DB_PATHS:
        if p and p.exists():
            return p
    # Search nearby
    for p in Path(".").glob("*.db"):
        return p
    return None


def clean(v):
    """Coerce openpyxl cell value to a clean string or None."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def clean_str(v):
    r = clean(v)
    if isinstance(r, float) and r == int(r):
        return str(int(r))
    return str(r) if r is not None else ""


def clean_float(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def clean_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def clean_date(v):
    """Accept YYYY-MM-DD string, datetime, or date object."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s   # return as-is, will fail gracefully on insert


def load_sheet(wb, name):
    """Return list of dicts from a sheet, skipping header+notes rows."""
    if name not in wb.sheetnames:
        print(f"  ⚠️  Sheet '{name}' not found in workbook — skipping.")
        return [], []

    ws = wb[name]
    headers = [clean_str(c.value) for c in ws[1]]   # row 1 = column names

    rows = []
    for row in ws.iter_rows(min_row=SKIP_ROWS + 1, values_only=True):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = row[i] if i < len(row) else None
        rows.append(d)
    return headers, rows


def print_section(title):
    print(f"\n{'─'*60}")
    print(f"  {title}")
    print(f"{'─'*60}")


# ─────────────────────────────────────────────────────────────────────────────
# Import functions
# ─────────────────────────────────────────────────────────────────────────────

def import_clients(conn, rows, dry_run, clear):
    print_section("Importing CLIENTS")
    if clear:
        conn.execute("UPDATE clients SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠️  Existing clients soft-deleted.")

    ok = skip = err = 0
    for r in rows:
        full_name = clean_str(r.get("full_name", ""))
        if not full_name:
            print(f"  ⚠️  Skipping row — full_name is required: {r}")
            skip += 1
            continue

        # Check for duplicate by full_name
        exists = conn.execute(
            "SELECT id FROM clients WHERE full_name=? AND is_deleted=0", [full_name]
        ).fetchone()
        if exists:
            print(f"  ↩  Skipping '{full_name}' — already exists (id={exists[0]})")
            skip += 1
            continue

        status = clean_str(r.get("status", "Active")) or "Active"
        if status not in ("Active", "Archived", "Prospect"):
            status = "Active"

        customer_id = clean_str(r.get("customer_id", ""))

        try:
            if not dry_run:
                conn.execute("""
                    INSERT INTO clients
                        (full_name, last_name, customer_id, year_acquired,
                         address, city_state_zip, phone1, phone2,
                         email1, email2, status, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, [
                    full_name,
                    clean_str(r.get("last_name", "")),
                    customer_id,
                    clean_int(r.get("year_acquired")) or None,
                    clean_str(r.get("address", "")),
                    clean_str(r.get("city_state_zip", "")),
                    clean_str(r.get("phone1", "")),
                    clean_str(r.get("phone2", "")),
                    clean_str(r.get("email1", "")),
                    clean_str(r.get("email2", "")),
                    status,
                    clean_str(r.get("notes", "")),
                ])
            print(f"  ✅  {'[DRY RUN] ' if dry_run else ''}Client: {full_name}")
            ok += 1
        except sqlite3.IntegrityError as e:
            print(f"  ❌  {full_name}: {e}")
            err += 1

    print(f"\n  Clients: {ok} imported, {skip} skipped, {err} errors")
    return ok, skip, err


def import_contractors(conn, rows, dry_run, clear):
    print_section("Importing CONTRACTORS")
    if clear:
        conn.execute("UPDATE contractors SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠️  Existing contractors soft-deleted.")

    ok = skip = err = 0
    for r in rows:
        company = clean_str(r.get("company_name", ""))
        if not company:
            print(f"  ⚠️  Skipping row — company_name is required: {r}")
            skip += 1
            continue

        exists = conn.execute(
            "SELECT id FROM contractors WHERE company_name=? AND is_deleted=0", [company]
        ).fetchone()
        if exists:
            print(f"  ↩  Skipping '{company}' — already exists (id={exists[0]})")
            skip += 1
            continue

        try:
            if not dry_run:
                conn.execute("""
                    INSERT INTO contractors
                        (company_name, trade_type, contact_person, phone, cell,
                         email, address, license_number, requires_1099,
                         rank_preference, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, [
                    company,
                    clean_str(r.get("trade_type", "")),
                    clean_str(r.get("contact_person", "")),
                    clean_str(r.get("phone", "")),
                    clean_str(r.get("cell", "")),
                    clean_str(r.get("email", "")),
                    clean_str(r.get("address", "")),
                    clean_str(r.get("license_number", "")),
                    clean_int(r.get("requires_1099", 0)),
                    clean_int(r.get("rank_preference", 0)),
                    clean_str(r.get("notes", "")),
                ])
            print(f"  ✅  {'[DRY RUN] ' if dry_run else ''}Contractor: {company}")
            ok += 1
        except sqlite3.IntegrityError as e:
            print(f"  ❌  {company}: {e}")
            err += 1

    print(f"\n  Contractors: {ok} imported, {skip} skipped, {err} errors")
    return ok, skip, err


def import_ledger(conn, rows, dry_run, clear):
    print_section("Importing LEDGER")
    if clear:
        conn.execute("UPDATE ledger SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠️  Existing ledger entries soft-deleted.")

    ok = skip = err = 0
    for i, r in enumerate(rows, 1):
        entry_date = clean_date(r.get("entry_date", ""))
        amount     = clean_float(r.get("amount", 0))
        vendor     = clean_str(r.get("vendor", ""))
        category   = clean_str(r.get("category", ""))

        if not entry_date:
            print(f"  ⚠️  Row {i}: missing entry_date — skipping")
            skip += 1
            continue
        if amount <= 0:
            print(f"  ⚠️  Row {i}: amount={amount} — skipping (must be > 0)")
            skip += 1
            continue
        if not vendor:
            print(f"  ⚠️  Row {i}: missing vendor — skipping")
            skip += 1
            continue

        is_cogs = clean_int(r.get("is_cogs", 0))
        job_code = clean_str(r.get("job_code", ""))
        # If job_code given but is_cogs not set, default is_cogs=1
        if job_code and is_cogs not in (0, 1):
            is_cogs = 1

        try:
            if not dry_run:
                conn.execute("""
                    INSERT INTO ledger
                        (entry_date, amount, vendor, category, description,
                         job_code, is_cogs, invoice_number,
                         receipt_filename, notes, status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,'Cleared')
                """, [
                    entry_date, amount, vendor, category,
                    clean_str(r.get("description", "")),
                    job_code, is_cogs,
                    clean_str(r.get("invoice_number", "")),
                    clean_str(r.get("receipt_filename", "")),
                    clean_str(r.get("notes", "")),
                ])
            print(f"  ✅  {'[DRY RUN] ' if dry_run else ''}{entry_date}  ${amount:>10,.2f}  {vendor[:24]:<24}  {category}")
            ok += 1
        except Exception as e:
            print(f"  ❌  Row {i} ({entry_date} {vendor}): {e}")
            err += 1

    print(f"\n  Ledger: {ok} imported, {skip} skipped, {err} errors")
    return ok, skip, err


def import_timesheet(conn, rows, dry_run, clear):
    print_section("Importing TIMESHEET")
    if clear:
        conn.execute("UPDATE timesheet SET is_deleted=1 WHERE is_deleted=0")
        print("  ⚠️  Existing timesheet entries soft-deleted.")

    ok = skip = err = 0
    for i, r in enumerate(rows, 1):
        entry_date   = clean_date(r.get("entry_date", ""))
        person_label = clean_str(r.get("person_label", ""))
        job_code     = clean_str(r.get("job_code", ""))
        hours        = clean_float(r.get("hours", 0))

        if not entry_date:
            print(f"  ⚠️  Row {i}: missing entry_date — skipping")
            skip += 1
            continue
        if not person_label:
            print(f"  ⚠️  Row {i}: missing person_label — skipping")
            skip += 1
            continue
        if hours <= 0:
            print(f"  ⚠️  Row {i}: hours={hours} — skipping")
            skip += 1
            continue

        cost_rate   = clean_float(r.get("cost_rate", 0))
        bill_rate   = clean_float(r.get("bill_rate", 0))
        cost_amount = round(hours * cost_rate, 2)
        bill_amount = round(hours * bill_rate, 2)

        try:
            if not dry_run:
                conn.execute("""
                    INSERT INTO timesheet
                        (entry_date, person_label, job_code, hours,
                         cost_rate, bill_rate, cost_amount, bill_amount,
                         description, invoice_number, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, [
                    entry_date, person_label, job_code, hours,
                    cost_rate, bill_rate, cost_amount, bill_amount,
                    clean_str(r.get("description", "")),
                    clean_str(r.get("invoice_number", "")),
                    clean_str(r.get("notes", "")),
                ])
            print(f"  ✅  {'[DRY RUN] ' if dry_run else ''}{entry_date}  {person_label:<18}  {job_code:<12}  {hours}h")
            ok += 1
        except Exception as e:
            print(f"  ❌  Row {i} ({entry_date} {person_label}): {e}")
            err += 1

    print(f"\n  Timesheet: {ok} imported, {skip} skipped, {err} errors")
    return ok, skip, err


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import data into Construction Manager")
    parser.add_argument("xlsx", help="Path to your filled-in import_data.xlsx")
    parser.add_argument("--db",             default=None, help="Path to construction.db (auto-detected if omitted)")
    parser.add_argument("--dry-run",        action="store_true", help="Preview what would be imported, make no changes")
    parser.add_argument("--clear-existing", action="store_true", help="Soft-delete existing records before importing")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"❌  File not found: {xlsx_path}")
        sys.exit(1)

    # Find database
    db_path = Path(args.db) if args.db else find_db()
    if not db_path or not db_path.exists():
        print("❌  Could not find the database file.")
        print("    Run the app once first (it creates the DB automatically).")
        print("    Or specify the path: --db path/to/construction.db")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  Construction Manager — Data Import")
    print(f"{'='*60}")
    print(f"  File  : {xlsx_path}")
    print(f"  DB    : {db_path}")
    print(f"  Mode  : {'DRY RUN (no changes will be made)' if args.dry_run else 'LIVE IMPORT'}")
    if args.clear_existing:
        print(f"  ⚠️   --clear-existing is ON — existing rows will be soft-deleted first!")
    print()

    if not args.dry_run and args.clear_existing:
        confirm = input("Type YES to confirm clearing existing data: ")
        if confirm.strip() != "YES":
            print("Aborted.")
            sys.exit(0)

    # Load workbook
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"❌  Could not open workbook: {e}")
        sys.exit(1)

    # Connect to DB
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = OFF")   # relax FK constraints during import

    totals = {"ok": 0, "skip": 0, "err": 0}

    def run(fn, sheet_name):
        _, rows = load_sheet(wb, sheet_name)
        print(f"\n  Found {len(rows)} data row(s) in '{sheet_name}' sheet.")
        o, s, e = fn(conn, rows, args.dry_run, args.clear_existing)
        totals["ok"]   += o
        totals["skip"] += s
        totals["err"]  += e

    run(import_clients,    "Clients")
    run(import_contractors,"Contractors")
    run(import_ledger,     "Ledger")
    run(import_timesheet,  "Timesheet")

    if not args.dry_run:
        conn.commit()

    conn.close()

    print(f"\n{'='*60}")
    print(f"  IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  ✅  Imported  : {totals['ok']}")
    print(f"  ↩  Skipped   : {totals['skip']}")
    print(f"  ❌  Errors    : {totals['err']}")
    if args.dry_run:
        print(f"\n  ℹ️  DRY RUN — nothing was written to the database.")
        print(f"      Remove --dry-run to perform the actual import.")
    else:
        print(f"\n  Start the app and verify your data looks correct.")
    print()


if __name__ == "__main__":
    main()
